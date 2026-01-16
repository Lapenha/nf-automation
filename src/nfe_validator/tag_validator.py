"""
Validador de tags XML obrigatórias baseado em planilha Excel.
Valida se todas as tags obrigatórias (1-1) estão presentes no XML.
"""

import logging
from pathlib import Path
from typing import List, Dict, Tuple, Optional
from dataclasses import dataclass
import openpyxl
from lxml import etree

logger = logging.getLogger(__name__)


@dataclass
class TagObrigatoria:
    """Representa uma tag obrigatória da planilha."""
    nome: str
    elemento: str
    ocorrencia: str
    linha: int
    tipo: str = "E"  # E = Elemento, G = Grupo, CE = Condicional, CG = Condicional Grupo
    descricao: str = ""
    observacao: str = ""


@dataclass
class ResultadoValidacao:
    """Resultado da validação de tags."""
    xml_path: Path
    total_tags_obrigatorias: int
    tags_encontradas: int
    tags_ausentes: List[TagObrigatoria]
    sucesso: bool
    mensagem: str


class TagValidator:
    """Validador de tags XML obrigatórias baseado em planilha Excel."""
    
    def __init__(self, excel_path: Path, aba_nome: str = "Emissão Nacional NFSe  - V1.01"):
        """
        Inicializa o validador.
        
        Args:
            excel_path: Caminho para a planilha Excel
            aba_nome: Nome da aba com as definições de tags
        """
        self.excel_path = excel_path
        self.aba_nome = aba_nome
        self.tags_obrigatorias: List[TagObrigatoria] = []
        self._carregar_tags_obrigatorias()
        
    def _carregar_tags_obrigatorias(self):
        """Carrega as tags obrigatórias da planilha Excel."""
        try:
            logger.info(f"Carregando tags obrigatórias de {self.excel_path}")
            wb = openpyxl.load_workbook(self.excel_path, data_only=True)
            
            if self.aba_nome not in wb.sheetnames:
                raise ValueError(f"Aba '{self.aba_nome}' não encontrada na planilha")
            
            ws = wb[self.aba_nome]
            
            # Percorre todas as linhas a partir da linha 9 (onde começam os dados)
            for i in range(9, ws.max_row + 1):
                nome = ws.cell(i, 10).value  # Coluna J - NOME DO CAMPO
                descricao = ws.cell(i, 11).value  # Coluna K - DESCRIÇÃO
                observacao = ws.cell(i, 12).value  # Coluna L - OBSERVAÇÃO
                elemento = ws.cell(i, 14).value  # Coluna N - ELEMENTO
                ocorrencia = ws.cell(i, 16).value  # Coluna P - OCORR.
                tipo = ws.cell(i, 15).value  # Coluna O - TIPO
                
                # Filtra apenas tags obrigatórias (1-1)
                if nome and elemento and ocorrencia and '1-1' in str(ocorrencia):
                    nome_limpo = nome.strip()
                    
                    tag = TagObrigatoria(
                        nome=nome_limpo,
                        elemento=str(elemento).strip(),
                        ocorrencia=str(ocorrencia).strip(),
                        linha=i,
                        tipo=str(tipo).strip() if tipo else "E",
                        descricao=str(descricao).strip() if descricao else "",
                        observacao=str(observacao).strip() if observacao else ""
                    )
                    self.tags_obrigatorias.append(tag)
            
            logger.info(f"Total de tags obrigatórias carregadas: {len(self.tags_obrigatorias)}")
            
        except Exception as e:
            logger.error(f"Erro ao carregar tags obrigatórias: {e}")
            raise
    
    def validar_xml(self, xml_path: Path) -> ResultadoValidacao:
        """
        Valida se um XML contém todas as tags obrigatórias.
        
        Args:
            xml_path: Caminho para o arquivo XML
            
        Returns:
            ResultadoValidacao com o resultado da validação
        """
        try:
            logger.info(f"Validando XML: {xml_path}")
            
            # Parse do XML
            tree = etree.parse(str(xml_path))
            root = tree.getroot()
            
            # Extrai todos os elementos do XML (sem namespace para simplificar)
            elementos_encontrados = set()
            self._extrair_elementos(root, elementos_encontrados)
            
            logger.debug(f"Total de elementos únicos encontrados no XML: {len(elementos_encontrados)}")
            
            # Verifica quais tags obrigatórias estão ausentes
            tags_ausentes = []
            for tag in self.tags_obrigatorias:
                # Verifica se o nome do campo está presente no XML
                if tag.nome not in elementos_encontrados:
                    tags_ausentes.append(tag)
            
            tags_encontradas = len(self.tags_obrigatorias) - len(tags_ausentes)
            sucesso = len(tags_ausentes) == 0
            
            if sucesso:
                mensagem = f"✅ Todas as {len(self.tags_obrigatorias)} tags obrigatórias estão presentes"
            else:
                mensagem = f"⚠️ {len(tags_ausentes)} tag(s) obrigatória(s) ausente(s) de {len(self.tags_obrigatorias)}"
            
            resultado = ResultadoValidacao(
                xml_path=xml_path,
                total_tags_obrigatorias=len(self.tags_obrigatorias),
                tags_encontradas=tags_encontradas,
                tags_ausentes=tags_ausentes,
                sucesso=sucesso,
                mensagem=mensagem
            )
            
            logger.info(f"Validação concluída: {mensagem}")
            return resultado
            
        except Exception as e:
            logger.error(f"Erro ao validar XML {xml_path}: {e}")
            return ResultadoValidacao(
                xml_path=xml_path,
                total_tags_obrigatorias=len(self.tags_obrigatorias),
                tags_encontradas=0,
                tags_ausentes=self.tags_obrigatorias,
                sucesso=False,
                mensagem=f"❌ Erro ao processar XML: {str(e)}"
            )
    
    def _extrair_elementos(self, element: etree._Element, elementos_set: set):
        """
        Extrai recursivamente todos os nomes de elementos do XML.
        
        Args:
            element: Elemento XML
            elementos_set: Set para armazenar os nomes dos elementos
        """
        # Remove namespace do tag
        tag_name = element.tag
        if '}' in tag_name:
            tag_name = tag_name.split('}')[1]
        
        elementos_set.add(tag_name)
        
        # Recursivamente processa os filhos
        for child in element:
            self._extrair_elementos(child, elementos_set)
    
    def validar_multiplos_xmls(self, xml_paths: List[Path]) -> List[ResultadoValidacao]:
        """
        Valida múltiplos arquivos XML.
        
        Args:
            xml_paths: Lista de caminhos para arquivos XML
            
        Returns:
            Lista de ResultadoValidacao
        """
        resultados = []
        for xml_path in xml_paths:
            resultado = self.validar_xml(xml_path)
            resultados.append(resultado)
        return resultados
    
    def gerar_relatorio_texto(self, resultados: List[ResultadoValidacao]) -> str:
        """
        Gera um relatório em texto dos resultados da validação.
        
        Args:
            resultados: Lista de resultados de validação
            
        Returns:
            String com o relatório
        """
        linhas = []
        linhas.append("=" * 80)
        linhas.append("RELATÓRIO DE VALIDAÇÃO DE TAGS OBRIGATÓRIAS")
        linhas.append("=" * 80)
        linhas.append("")
        
        total_xmls = len(resultados)
        total_sucesso = sum(1 for r in resultados if r.sucesso)
        total_falha = total_xmls - total_sucesso
        
        linhas.append(f"📊 RESUMO GERAL")
        linhas.append(f"   Total de XMLs validados: {total_xmls}")
        linhas.append(f"   ✅ XMLs válidos: {total_sucesso}")
        linhas.append(f"   ❌ XMLs com tags ausentes: {total_falha}")
        linhas.append("")
        
        linhas.append("=" * 80)
        linhas.append("DETALHES POR ARQUIVO")
        linhas.append("=" * 80)
        linhas.append("")
        
        for idx, resultado in enumerate(resultados, 1):
            linhas.append(f"{idx}. {resultado.xml_path.name}")
            linhas.append(f"   {resultado.mensagem}")
            linhas.append(f"   Tags encontradas: {resultado.tags_encontradas}/{resultado.total_tags_obrigatorias}")
            
            if not resultado.sucesso and resultado.tags_ausentes:
                linhas.append(f"   Tags ausentes ({len(resultado.tags_ausentes)}):")
                for tag in resultado.tags_ausentes[:20]:  # Limita a 20 para não ficar muito grande
                    linhas.append(f"      - {tag.nome} ({tag.elemento}) - Linha {tag.linha}")
                
                if len(resultado.tags_ausentes) > 20:
                    linhas.append(f"      ... e mais {len(resultado.tags_ausentes) - 20} tag(s)")
            
            linhas.append("")
        
        return "\n".join(linhas)


def main():
    """Função de teste."""
    import sys
    
    # Configura logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # Caminho da planilha
    base_dir = Path(__file__).parent.parent.parent
    excel_path = base_dir / "planilhabase" / "Mastersaf_Layout_DFe_V3_NFSe_NACIONAL_1_01 - Oficial Reforma.xlsx"
    
    if not excel_path.exists():
        print(f"❌ Planilha não encontrada: {excel_path}")
        sys.exit(1)
    
    # Cria validador
    validator = TagValidator(excel_path)
    print(f"✅ Validador criado com {len(validator.tags_obrigatorias)} tags obrigatórias")
    
    # Se passou XMLs como argumento, valida
    if len(sys.argv) > 1:
        xml_paths = [Path(arg) for arg in sys.argv[1:]]
        resultados = validator.validar_multiplos_xmls(xml_paths)
        
        # Gera relatório
        relatorio = validator.gerar_relatorio_texto(resultados)
        print("\n")
        print(relatorio)
    else:
        print("\nPara validar XMLs, execute:")
        print("  python -m nfe_validator.tag_validator <arquivo1.xml> <arquivo2.xml> ...")


if __name__ == "__main__":
    main()
