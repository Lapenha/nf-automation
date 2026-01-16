"""
Interface gráfica para validação de NF-e (IBS/CBS).
Interface moderna com drag-and-drop, seleção de pasta e barra de progresso.
"""
import sys
import os
import ctypes
from pathlib import Path
from typing import List, Optional
from datetime import datetime
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QProgressBar, QFileDialog, QMessageBox,
    QTextEdit, QFrame, QScrollArea
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QMimeData
from PyQt6.QtGui import QDragEnterEvent, QDropEvent, QFont, QIcon

from .cli import NFeValidator
from .config import load_config, Config
from .tag_validator import TagValidator, ResultadoValidacao


class ValidationThread(QThread):
    """Thread para processar validação sem travar a interface."""
    
    progress = pyqtSignal(int, int)  # current, total
    log_message = pyqtSignal(str)
    finished = pyqtSignal(str, bool)  # output_path, success
    
    def __init__(self, xml_paths: List[Path], output_dir: Path):
        super().__init__()
        self.xml_paths = xml_paths
        self.output_dir = output_dir
        self._is_running = True
        
    def run(self):
        """Executa a validação em background."""
        print("🔄 Thread iniciou execução!")
        self.log_message.emit("🔄 Thread iniciou execução!")
        try:
            print("🚀 Iniciando validação...")
            self.log_message.emit("🚀 Iniciando validação...")
            print(f"📁 {len(self.xml_paths)} arquivo(s) XML encontrado(s)")
            self.log_message.emit(f"📁 {len(self.xml_paths)} arquivo(s) XML encontrado(s)\n")
            
            # Carrega configuração (usa padrão se não encontrar)
            try:
                print("⚙️ Carregando configuração...")
                self.log_message.emit("⚙️  Carregando configuração...\n")
                config = Config()  # Usa config padrão sempre por enquanto
                print("✅ Configuração criada")
                self.log_message.emit("✅ Configuração padrão carregada\n")
            except Exception as e:
                print(f"❌ Erro ao criar config: {e}")
                self.log_message.emit(f"❌ Erro ao criar config: {e}\n")
                raise
            
            # Cria validador
            print("🔧 Criando validador...")
            self.log_message.emit("🔧 Criando validador...\n")
            validator = NFeValidator(config)
            print("✅ Validador criado!")
            self.log_message.emit("✅ Validador criado!\n")
            
            # Processa XMLs
            total = len(self.xml_paths)
            print(f"📝 Total de arquivos: {total}")
            self.log_message.emit(f"📝 Total de arquivos: {total}\n")
            
            for idx, xml_path in enumerate(self.xml_paths, 1):
                if not self._is_running:
                    print("❌ Processamento cancelado pelo usuário")
                    self.log_message.emit("\n❌ Processamento cancelado pelo usuário")
                    return
                    
                print(f"⚙️ [{idx}/{total}] Processando: {xml_path.name}")
                self.log_message.emit(f"\n⚙️  [{idx}/{total}] Processando: {xml_path.name}")
                self.log_message.emit(f"    Caminho: {xml_path}")
                self.progress.emit(idx, total)
                
                try:
                    print(f"  → Fazendo parse do XML...")
                    self.log_message.emit(f"    → Fazendo parse do XML...")
                    validator.process_file(xml_path)
                    print(f"  ✅ OK!")
                    self.log_message.emit(f"    ✅ OK!")
                except Exception as e:
                    import traceback
                    print(f"  ❌ Erro: {str(e)}")
                    print(f"  Stack: {traceback.format_exc()}")
                    self.log_message.emit(f"    ❌ Erro: {str(e)}")
                    self.log_message.emit(f"    Stack: {traceback.format_exc()}")
            
            # Gera relatório
            print("\n📊 Gerando relatório Excel...")
            self.log_message.emit("\n📊 Gerando relatório Excel...")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            report_name = f"relatorio_ibs_cbs_{timestamp}.xlsx"
            report_path = self.output_dir / report_name
            
            print(f"📁 Caminho do relatório: {report_path}")
            self.log_message.emit(f"📁 Caminho do relatório: {report_path}")
            
            # Cria diretório se não existir
            try:
                report_path.parent.mkdir(parents=True, exist_ok=True)
                print("✅ Diretório criado/verificado")
                self.log_message.emit(f"✅ Diretório criado/verificado")
            except Exception as e:
                print(f"⚠️ Erro ao criar diretório: {e}")
                self.log_message.emit(f"⚠️  Erro ao criar diretório: {e}")
            
            try:
                print("📝 Chamando generate_report()...")
                validator.generate_report(report_path)
                print("✅ Relatório gerado com sucesso!")
                self.log_message.emit("✅ Relatório gerado com sucesso!")
            except Exception as e:
                import traceback
                print(f"❌ Erro ao gerar relatório: {e}")
                print(f"🔍 Stack trace:\n{traceback.format_exc()}")
                self.log_message.emit(f"❌ Erro ao gerar relatório: {e}")
                self.log_message.emit(f"\n🔍 Stack trace:\n{traceback.format_exc()}")
                raise
            
            # Estatísticas
            stats = validator.get_statistics()
            print("="*50)
            print("📈 RESUMO DA VALIDAÇÃO")
            print("="*50)
            print(f"✅ NF-es processadas: {stats['total_nfes']}")
            print(f"📦 Total de itens: {stats['total_items']}")
            print(f"✔️ Itens OK: {stats['items_ok']}")
            self.log_message.emit("\n" + "="*50)
            self.log_message.emit("📈 RESUMO DA VALIDAÇÃO")
            self.log_message.emit("="*50)
            self.log_message.emit(f"✅ NF-es processadas: {stats['total_nfes']}")
            self.log_message.emit(f"📦 Total de itens: {stats['total_items']}")
            self.log_message.emit(f"✔️  Itens OK: {stats['items_ok']}")
            print(f"⚠️ Divergências IBS: {stats['divergencias_ibs']}")
            print(f"⚠️ Divergências CBS: {stats['divergencias_cbs']}")
            print(f"❌ Erros: {stats['total_errors']}")
            print("="*50)
            self.log_message.emit(f"⚠️  Divergências IBS: {stats['divergencias_ibs']}")
            self.log_message.emit(f"⚠️  Divergências CBS: {stats['divergencias_cbs']}")
            self.log_message.emit(f"❌ Erros: {stats['total_errors']}")
            self.log_message.emit("="*50 + "\n")
            
            print(f"✨ Relatório salvo em: {report_path}")
            print("🎉 Validação concluída!")
            self.log_message.emit(f"✨ Relatório salvo em:\n   {report_path}\n")
            self.finished.emit(str(report_path), True)
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"\n❌ ERRO: {str(e)}")
            print(f"\n🔍 Detalhes técnicos:\n{error_details}")
            self.log_message.emit(f"\n❌ ERRO: {str(e)}")
            self.log_message.emit(f"\n🔍 Detalhes técnicos:\n{error_details}")
            self.finished.emit("", False)
    
    def stop(self):
        """Para a thread."""
        self._is_running = False


class TagValidationThread(QThread):
    """Thread para validação de tags obrigatórias sem travar a interface."""
    
    progress = pyqtSignal(int, int)  # current, total
    log_message = pyqtSignal(str)
    finished = pyqtSignal(str, bool)  # output_path, success
    
    def __init__(self, xml_paths: List[Path], output_dir: Path, excel_path: Path):
        super().__init__()
        self.xml_paths = xml_paths
        self.output_dir = output_dir
        self.excel_path = excel_path
        self._is_running = True
        
    def run(self):
        """Executa a validação de tags em background."""
        try:
            self.log_message.emit("🔄 Iniciando validação de tags obrigatórias...")
            self.log_message.emit(f"📁 {len(self.xml_paths)} arquivo(s) XML encontrado(s)\n")
            
            # Cria validador
            self.log_message.emit("⚙️  Carregando planilha de referência...")
            validator = TagValidator(self.excel_path)
            self.log_message.emit(f"✅ {len(validator.tags_obrigatorias)} tags obrigatórias carregadas\n")
            
            # Valida XMLs
            resultados = []
            total = len(self.xml_paths)
            
            for idx, xml_path in enumerate(self.xml_paths, 1):
                if not self._is_running:
                    self.log_message.emit("\n❌ Validação cancelada pelo usuário")
                    return
                    
                self.log_message.emit(f"\n⚙️  [{idx}/{total}] Validando: {xml_path.name}")
                self.progress.emit(idx, total)
                
                try:
                    resultado = validator.validar_xml(xml_path)
                    resultados.append(resultado)
                    
                    if resultado.sucesso:
                        self.log_message.emit(f"    ✅ Todas as tags obrigatórias presentes")
                    else:
                        self.log_message.emit(f"    ⚠️  {len(resultado.tags_ausentes)} tag(s) ausente(s)")
                        
                except Exception as e:
                    self.log_message.emit(f"    ❌ Erro: {str(e)}")
            
            # Gera relatório Excel
            self.log_message.emit("\n📊 Gerando relatório Excel...")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            report_name = f"relatorio_tags_obrigatorias_{timestamp}.xlsx"
            report_path = self.output_dir / report_name
            
            self._gerar_relatorio_excel(resultados, report_path)
            
            # Estatísticas
            total_sucesso = sum(1 for r in resultados if r.sucesso)
            total_falha = len(resultados) - total_sucesso
            
            self.log_message.emit("\n" + "="*50)
            self.log_message.emit("📈 RESUMO DA VALIDAÇÃO DE TAGS")
            self.log_message.emit("="*50)
            self.log_message.emit(f"✅ XMLs válidos: {total_sucesso}")
            self.log_message.emit(f"⚠️  XMLs com tags ausentes: {total_falha}")
            self.log_message.emit(f"📝 Total de XMLs: {len(resultados)}")
            self.log_message.emit("="*50 + "\n")
            
            self.log_message.emit(f"✨ Relatório salvo em:\n   {report_path}\n")
            self.finished.emit(str(report_path), True)
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            self.log_message.emit(f"\n❌ ERRO: {str(e)}")
            self.log_message.emit(f"\n🔍 Detalhes técnicos:\n{error_details}")
            self.finished.emit("", False)
    
    def _gerar_relatorio_excel(self, resultados: List[ResultadoValidacao], output_path: Path):
        """Gera relatório Excel com os resultados da validação."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            
            # Aba 1: Resumo
            ws_resumo = wb.active
            ws_resumo.title = "Resumo"
            
            # Cabeçalho
            ws_resumo['A1'] = "RELATÓRIO DE VALIDAÇÃO DE TAGS OBRIGATÓRIAS"
            ws_resumo['A1'].font = Font(size=14, bold=True)
            ws_resumo.merge_cells('A1:F1')
            
            ws_resumo['A3'] = "Arquivo XML"
            ws_resumo['B3'] = "Status"
            ws_resumo['C3'] = "Tags Encontradas"
            ws_resumo['D3'] = "Total Tags"
            ws_resumo['E3'] = "Tags Ausentes"
            ws_resumo['F3'] = "% Completude"
            
            for cell in ws_resumo[3]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            # Dados
            linha = 4
            for resultado in resultados:
                ws_resumo[f'A{linha}'] = resultado.xml_path.name
                ws_resumo[f'B{linha}'] = "✅ OK" if resultado.sucesso else "⚠️ Ausências"
                ws_resumo[f'C{linha}'] = resultado.tags_encontradas
                ws_resumo[f'D{linha}'] = resultado.total_tags_obrigatorias
                ws_resumo[f'E{linha}'] = len(resultado.tags_ausentes)
                
                percentual = (resultado.tags_encontradas / resultado.total_tags_obrigatorias * 100) if resultado.total_tags_obrigatorias > 0 else 0
                ws_resumo[f'F{linha}'] = f"{percentual:.1f}%"
                
                # Colorir linha baseado no status
                if resultado.sucesso:
                    fill_color = "C6EFCE"  # Verde claro
                else:
                    fill_color = "FFEB9C"  # Amarelo claro
                
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    ws_resumo[f'{col}{linha}'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                
                linha += 1
            
            # Ajusta largura das colunas
            ws_resumo.column_dimensions['A'].width = 50
            ws_resumo.column_dimensions['B'].width = 15
            ws_resumo.column_dimensions['C'].width = 18
            ws_resumo.column_dimensions['D'].width = 15
            ws_resumo.column_dimensions['E'].width = 18
            ws_resumo.column_dimensions['F'].width = 15
            
            # Aba 2: Detalhes das tags ausentes
            ws_detalhes = wb.create_sheet("Tags Ausentes")
            
            ws_detalhes['A1'] = "TAGS OBRIGATÓRIAS AUSENTES POR ARQUIVO"
            ws_detalhes['A1'].font = Font(size=14, bold=True)
            ws_detalhes.merge_cells('A1:G1')
            
            ws_detalhes['A3'] = "Arquivo XML"
            ws_detalhes['B3'] = "Nome da Tag"
            ws_detalhes['C3'] = "Descrição"
            ws_detalhes['D3'] = "Observação"
            ws_detalhes['E3'] = "Elemento"
            ws_detalhes['F3'] = "Tipo"
            ws_detalhes['G3'] = "Linha Planilha"
            
            for cell in ws_detalhes[3]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            linha = 4
            for resultado in resultados:
                if resultado.tags_ausentes:
                    for tag in resultado.tags_ausentes:
                        ws_detalhes[f'A{linha}'] = resultado.xml_path.name
                        ws_detalhes[f'B{linha}'] = tag.nome
                        ws_detalhes[f'C{linha}'] = tag.descricao
                        ws_detalhes[f'D{linha}'] = tag.observacao
                        ws_detalhes[f'E{linha}'] = tag.elemento
                        ws_detalhes[f'F{linha}'] = tag.tipo
                        ws_detalhes[f'G{linha}'] = tag.linha
                        linha += 1
            
            # Ajusta largura das colunas
            ws_detalhes.column_dimensions['A'].width = 50
            ws_detalhes.column_dimensions['B'].width = 30
            ws_detalhes.column_dimensions['C'].width = 50
            ws_detalhes.column_dimensions['D'].width = 50
            ws_detalhes.column_dimensions['E'].width = 15
            ws_detalhes.column_dimensions['F'].width = 10
            ws_detalhes.column_dimensions['G'].width = 18
            
            # Salva
            wb.save(output_path)
            self.log_message.emit(f"✅ Relatório Excel gerado com sucesso")
            
        except Exception as e:
            self.log_message.emit(f"❌ Erro ao gerar relatório Excel: {e}")
            raise
    
    def stop(self):
        """Para a thread."""
        self._is_running = False


class DropArea(QFrame):
    """Área para arrastar e soltar arquivos XML."""
    
    files_dropped = pyqtSignal(list)
    
    def __init__(self):
        super().__init__()
        self.setAcceptDrops(True)
        self.setFrameStyle(QFrame.Shape.StyledPanel | QFrame.Shadow.Sunken)
        self.setStyleSheet("""
            DropArea {
                background-color: #f8f9fa;
                border: 2px dashed #007bff;
                border-radius: 10px;
                padding: 40px;
            }
            DropArea:hover {
                background-color: #e9ecef;
                border-color: #0056b3;
            }
        """)
        
        layout = QVBoxLayout()
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # Ícone de upload (emoji)
        icon_label = QLabel("📁")
        icon_label.setStyleSheet("font-size: 64px; border: none;")
        icon_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # Texto
        text_label = QLabel("Arraste arquivos XML aqui\nou clique em 'Selecionar Pasta'")
        text_label.setStyleSheet("font-size: 16px; color: #6c757d; border: none;")
        text_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        layout.addWidget(icon_label)
        layout.addWidget(text_label)
        self.setLayout(layout)
        
    def dragEnterEvent(self, event: QDragEnterEvent):
        """Aceita arquivos XML arrastados."""
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            
    def dropEvent(self, event: QDropEvent):
        """Processa arquivos soltos."""
        files = []
        for url in event.mimeData().urls():
            file_path = Path(url.toLocalFile())
            if file_path.suffix.lower() == '.xml':
                files.append(file_path)
                
        if files:
            self.files_dropped.emit(files)


class MainWindow(QMainWindow):
    """Janela principal da aplicação."""
    
    def __init__(self):
        super().__init__()
        self.xml_files: List[Path] = []
        self.validation_thread: Optional[ValidationThread] = None
        self.tag_validation_thread: Optional[TagValidationThread] = None
        self.output_dir = Path.home() / "Downloads"
        
        # Caminho da planilha de referência
        base_dir = Path(__file__).parent.parent.parent
        self.excel_path = base_dir / "planilhabase" / "Mastersaf_Layout_DFe_V3_NFSe_NACIONAL_1_01 - Oficial Reforma.xlsx"
        
        self.init_ui()
        
    def init_ui(self):
        """Inicializa a interface."""
        self.setWindowTitle("Validador NF-e - IBS/CBS (Reforma Tributária)")
        self.setGeometry(100, 100, 900, 700)
        
        # Widget central com scroll
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setStyleSheet("""
            QScrollArea {
                border: none;
                background-color: white;
            }
        """)
        self.setCentralWidget(scroll_area)
        
        # Container para o conteúdo
        central_widget = QWidget()
        scroll_area.setWidget(central_widget)
        
        # Layout principal
        main_layout = QVBoxLayout()
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(30, 30, 30, 30)
        
        # Título
        title = QLabel("🧾 Validador de NF-e - IBS/CBS")
        title.setStyleSheet("""
            font-size: 28px;
            font-weight: bold;
            color: black;
            padding: 10px;
        """)
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(title)
        
        # Subtítulo
        subtitle = QLabel("Validação automática de tributos da Reforma Tributária")
        subtitle.setStyleSheet("font-size: 14px; color: black;")
        subtitle.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(subtitle)
        
        # Área de drop
        self.drop_area = DropArea()
        self.drop_area.files_dropped.connect(self.add_xml_files)
        main_layout.addWidget(self.drop_area)
        
        # Botões
        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)
        
        self.folder_btn = QPushButton("📂 Selecionar Pasta")
        self.folder_btn.setStyleSheet(self.get_button_style("#007bff"))
        self.folder_btn.clicked.connect(self.select_folder)
        self.folder_btn.setMinimumHeight(50)
        
        self.file_btn = QPushButton("📄 Adicionar Arquivo(s)")
        self.file_btn.setStyleSheet(self.get_button_style("#17a2b8"))
        self.file_btn.clicked.connect(self.select_files)
        self.file_btn.setMinimumHeight(50)
        
        self.clear_btn = QPushButton("🗑️ Limpar")
        self.clear_btn.setStyleSheet(self.get_button_style("#6c757d"))
        self.clear_btn.clicked.connect(self.clear_files)
        self.clear_btn.setMinimumHeight(50)
        
        button_layout.addWidget(self.folder_btn)
        button_layout.addWidget(self.file_btn)
        button_layout.addWidget(self.clear_btn)
        main_layout.addLayout(button_layout)
        
        # Label de arquivos selecionados
        self.files_label = QLabel("Nenhum arquivo selecionado")
        self.files_label.setStyleSheet("font-size: 13px; color: black; padding: 10px;")
        self.files_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(self.files_label)
        
        # Botão de processar
        self.process_btn = QPushButton("▶️ INICIAR VALIDAÇÃO IBS/CBS")
        self.process_btn.setStyleSheet(self.get_button_style("#28a745", 60))
        self.process_btn.clicked.connect(self.start_validation)
        self.process_btn.setEnabled(False)
        self.process_btn.setMinimumHeight(60)
        main_layout.addWidget(self.process_btn)
        
        # Botão de validar tags XML
        self.tag_validation_btn = QPushButton("🏷️ VALIDAR TAGS XML OBRIGATÓRIAS")
        self.tag_validation_btn.setStyleSheet(self.get_button_style("#17a2b8", 60))
        self.tag_validation_btn.clicked.connect(self.start_tag_validation)
        self.tag_validation_btn.setEnabled(False)
        self.tag_validation_btn.setMinimumHeight(60)
        main_layout.addWidget(self.tag_validation_btn)
        
        # Barra de progresso
        self.progress_bar = QProgressBar()
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 2px solid #dee2e6;
                border-radius: 5px;
                text-align: center;
                font-size: 14px;
                font-weight: bold;
                height: 30px;
            }
            QProgressBar::chunk {
                background-color: #28a745;
                border-radius: 3px;
            }
        """)
        self.progress_bar.setVisible(False)
        main_layout.addWidget(self.progress_bar)
        
        # Log de processamento
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setStyleSheet("""
            QTextEdit {
                background-color: white;
                color: black;
                border: 1px solid #dee2e6;
                border-radius: 5px;
                padding: 10px;
                font-family: 'Consolas', 'Courier New', monospace;
                font-size: 11px;
            }
        """)
        self.log_text.setVisible(True)  # Sempre visível
        self.log_text.setMinimumHeight(150)
        self.log_text.setMaximumHeight(250)
        self.log_text.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.log_text.append("📋 Aguardando arquivos XML...")
        main_layout.addWidget(self.log_text)
        
        central_widget.setLayout(main_layout)
        
        # Estilo geral da janela
        self.setStyleSheet("""
            QMainWindow {
                background-color: white;
            }
            QWidget {
                background-color: white;
                color: black;
            }
        """)
    
    def get_button_style(self, color: str, height: int = 50) -> str:
        """Retorna o estilo CSS para botões."""
        return f"""
            QPushButton {{
                background-color: {color};
                color: white;
                border: none;
                border-radius: 8px;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: bold;
                min-height: {height}px;
            }}
            QPushButton:hover {{
                background-color: {self.darken_color(color)};
            }}
            QPushButton:pressed {{
                background-color: {self.darken_color(color, 0.3)};
            }}
            QPushButton:disabled {{
                background-color: #e9ecef;
                color: #6c757d;
            }}
        """
    
    def darken_color(self, hex_color: str, factor: float = 0.15) -> str:
        """Escurece uma cor hex."""
        hex_color = hex_color.lstrip('#')
        r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
        r, g, b = int(r * (1 - factor)), int(g * (1 - factor)), int(b * (1 - factor))
        return f"#{r:02x}{g:02x}{b:02x}"
    
    def select_folder(self):
        """Seleciona pasta com arquivos XML."""
        folder = QFileDialog.getExistingDirectory(self, "Selecionar Pasta com XMLs")
        if folder:
            xml_files = list(Path(folder).glob("*.xml"))
            if xml_files:
                self.add_xml_files(xml_files)
            else:
                QMessageBox.warning(self, "Aviso", "Nenhum arquivo XML encontrado na pasta selecionada.")
    
    def select_files(self):
        """Seleciona arquivos XML individuais."""
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Selecionar Arquivos XML",
            "",
            "Arquivos XML (*.xml)"
        )
        if files:
            self.add_xml_files([Path(f) for f in files])
    
    def add_xml_files(self, files: List[Path]):
        """Adiciona arquivos XML à lista."""
        for file in files:
            if file not in self.xml_files:
                self.xml_files.append(file)
        
        self.update_files_label()
        self.process_btn.setEnabled(len(self.xml_files) > 0)
        self.tag_validation_btn.setEnabled(len(self.xml_files) > 0)
    
    def clear_files(self):
        """Limpa lista de arquivos."""
        self.xml_files.clear()
        self.update_files_label()
        self.process_btn.setEnabled(False)
        self.tag_validation_btn.setEnabled(False)
        self.log_text.clear()
        self.log_text.setVisible(False)
        self.progress_bar.setVisible(False)
    
    def update_files_label(self):
        """Atualiza label com quantidade de arquivos."""
        count = len(self.xml_files)
        if count == 0:
            self.files_label.setText("Nenhum arquivo selecionado")
        elif count == 1:
            self.files_label.setText(f"1 arquivo selecionado: {self.xml_files[0].name}")
        else:
            self.files_label.setText(f"{count} arquivos selecionados")
    
    def start_validation(self):
        """Inicia o processamento."""
        if not self.xml_files:
            self.log_text.append("❌ Nenhum arquivo selecionado!")
            return
        
        self.log_text.append(f"🚀 Iniciando validação de {len(self.xml_files)} arquivo(s)...")
        
        # Configura interface
        self.process_btn.setEnabled(False)
        self.folder_btn.setEnabled(False)
        self.file_btn.setEnabled(False)
        self.clear_btn.setEnabled(False)
        
        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(True)
        
        self.log_text.append(f"📁 Pasta de saída: {self.output_dir}")
        
        # Cria e inicia thread
        try:
            self.validation_thread = ValidationThread(self.xml_files, self.output_dir)
            self.validation_thread.progress.connect(self.update_progress)
            self.validation_thread.log_message.connect(self.add_log_message)
            self.validation_thread.finished.connect(self.validation_finished)
            self.log_text.append("✅ Thread criada, iniciando...")
            self.validation_thread.start()
            self.log_text.append("✅ Thread iniciada!")
        except Exception as e:
            self.log_text.append(f"❌ Erro ao criar thread: {e}")
            import traceback
            self.log_text.append(traceback.format_exc())
    
    def start_tag_validation(self):
        """Inicia a validação de tags obrigatórias."""
        if not self.xml_files:
            self.log_text.append("❌ Nenhum arquivo selecionado!")
            return
        
        # Verifica se a planilha existe
        if not self.excel_path.exists():
            QMessageBox.critical(
                self,
                "Erro",
                f"❌ Planilha de referência não encontrada:\n{self.excel_path}\n\n"
                "Por favor, certifique-se de que a planilha está no local correto."
            )
            return
        
        self.log_text.append(f"🏷️ Iniciando validação de tags de {len(self.xml_files)} arquivo(s)...")
        
        # Configura interface
        self.process_btn.setEnabled(False)
        self.tag_validation_btn.setEnabled(False)
        self.folder_btn.setEnabled(False)
        self.file_btn.setEnabled(False)
        self.clear_btn.setEnabled(False)
        
        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(True)
        
        self.log_text.append(f"📁 Pasta de saída: {self.output_dir}")
        self.log_text.append(f"📋 Planilha: {self.excel_path.name}")
        
        # Cria e inicia thread
        try:
            self.tag_validation_thread = TagValidationThread(
                self.xml_files, 
                self.output_dir,
                self.excel_path
            )
            self.tag_validation_thread.progress.connect(self.update_progress)
            self.tag_validation_thread.log_message.connect(self.add_log_message)
            self.tag_validation_thread.finished.connect(self.validation_finished)
            self.log_text.append("✅ Thread de validação criada, iniciando...")
            self.tag_validation_thread.start()
            self.log_text.append("✅ Thread iniciada!")
        except Exception as e:
            self.log_text.append(f"❌ Erro ao criar thread: {e}")
            import traceback
            self.log_text.append(traceback.format_exc())
    
    def update_progress(self, current: int, total: int):
        """Atualiza barra de progresso."""
        percentage = int((current / total) * 100)
        self.progress_bar.setValue(percentage)
        self.progress_bar.setFormat(f"{current}/{total} arquivos - {percentage}%")
    
    def add_log_message(self, message: str):
        """Adiciona mensagem ao log."""
        self.log_text.append(message)
        self.log_text.ensureCursorVisible()
        self.log_text.verticalScrollBar().setValue(
            self.log_text.verticalScrollBar().maximum()
        )
        # Force update
        self.log_text.repaint()
    
    def validation_finished(self, output_path: str, success: bool):
        """Chamado quando a validação termina."""
        # Reabilita interface
        self.process_btn.setEnabled(True)
        self.tag_validation_btn.setEnabled(True)
        self.folder_btn.setEnabled(True)
        self.file_btn.setEnabled(True)
        self.clear_btn.setEnabled(True)
        
        if success and output_path:
            # Mensagem de sucesso
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Icon.Information)
            msg.setWindowTitle("Validação Concluída")
            msg.setText("✅ Validação concluída com sucesso!")
            msg.setInformativeText(f"Relatório salvo em Downloads:\n{Path(output_path).name}")
            
            # Botões
            open_btn = msg.addButton("Abrir Relatório", QMessageBox.ButtonRole.AcceptRole)
            open_folder_btn = msg.addButton("Abrir Pasta", QMessageBox.ButtonRole.ActionRole)
            close_btn = msg.addButton("Fechar", QMessageBox.ButtonRole.RejectRole)
            
            result = msg.exec()
            clicked = msg.clickedButton()
            
            # Ações
            if clicked == open_btn:
                try:
                    os.startfile(output_path)
                except Exception as e:
                    QMessageBox.warning(self, "Aviso", f"Não foi possível abrir o arquivo:\n{e}")
            elif clicked == open_folder_btn:
                try:
                    os.startfile(str(Path(output_path).parent))
                except Exception as e:
                    QMessageBox.warning(self, "Aviso", f"Não foi possível abrir a pasta:\n{e}")
        else:
            QMessageBox.critical(
                self,
                "Erro",
                "❌ Ocorreu um erro durante a validação.\nVerifique o log para mais detalhes."
            )


def main():
    """Função principal da GUI."""
    # On Windows, set an explicit AppUserModelID so the shell uses the
    # application's icon (and groups windows correctly). This should be
    # done before creating windows / the QApplication when possible.
    if sys.platform.startswith("win"):
        try:
            app_id = u"com.lapenha.validadornfe"
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(app_id)
        except Exception:
            # Not fatal — continue even if API call fails
            pass

    app = QApplication(sys.argv)

    # Define o ícone da aplicação para garantir que apareça em todos os lugares
    def _resource_path(name: str) -> str:
        # Quando empacotado com PyInstaller, recursos ficam em sys._MEIPASS
        if getattr(sys, 'frozen', False):
            base = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
        else:
            # Ao rodar a partir do código-fonte, procura no diretório do projeto
            base = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
        candidate = os.path.join(base, name)
        if os.path.exists(candidate):
            return candidate
        # fallback: procurar no cwd
        if os.path.exists(name):
            return os.path.abspath(name)
        return candidate

    icon_path = _resource_path('icon.ico')
    if not os.path.exists(icon_path):
        # tenta png como fallback
        icon_path = _resource_path('icon.png')
    try:
        if os.path.exists(icon_path):
            app.setWindowIcon(QIcon(icon_path))
    except Exception:
        pass
    
    # Fonte padrão
    font = QFont("Segoe UI", 10)
    app.setFont(font)
    
    # Janela principal
    window = MainWindow()
    # garante que a janela principal também use o mesmo ícone
    try:
        if os.path.exists(icon_path):
            window.setWindowIcon(QIcon(icon_path))
    except Exception:
        pass
    window.show()
    
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
