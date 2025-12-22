"""
Gerador de relatórios Excel com formatação para auditoria.
"""

import logging
from pathlib import Path
from datetime import datetime
from typing import List
from decimal import Decimal

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import ComparisonResult, ErrorRecord, ValidationStatus

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Gera relatório Excel consolidado."""
    
    def __init__(self, output_path: Path):
        """
        Inicializa gerador de relatórios.
        
        Args:
            output_path: Caminho do arquivo Excel de saída
        """
        self.output_path = output_path
    
    def generate(
        self,
        results: List[ComparisonResult],
        errors: List[ErrorRecord],
        config_summary: dict = None,
    ):
        """
        Gera relatório Excel completo.
        
        Args:
            results: Lista de resultados de comparação
            errors: Lista de erros de processamento
            config_summary: Resumo da configuração usada (opcional)
        """
        logger.info(f"Gerando relatório Excel: {self.output_path}")
        
        # Criar DataFrames
        df_resumo = self._create_resumo_df(results)
        df_itens = self._create_itens_df(results)
        df_divergencias = self._create_divergencias_df(results)
        df_erros = self._create_erros_df(errors)
        
        # Escrever Excel
        with pd.ExcelWriter(self.output_path, engine="openpyxl") as writer:
            df_resumo.to_excel(writer, sheet_name="Resumo NF", index=False)
            df_itens.to_excel(writer, sheet_name="Itens", index=False)
            df_divergencias.to_excel(writer, sheet_name="Divergências", index=False)
            df_erros.to_excel(writer, sheet_name="Erros", index=False)
            # Aba com todas as tags extraídas (por nota)
            df_tags = self._create_tags_df(results)
            df_tags.to_excel(writer, sheet_name="Tags", index=False)
            # Aba combinada: todos os dados concatenados por documento com separadores
            self._create_combined_sheet(writer, results)
            
            # Adicionar aba de configuração se fornecida
            if config_summary:
                df_config = pd.DataFrame([config_summary]).T
                df_config.columns = ["Valor"]
                df_config.to_excel(writer, sheet_name="Configuração", header=True)
        
        # Aplicar formatação
        self._apply_formatting()
        
        logger.info(f"Relatório gerado com sucesso: {self.output_path}")
    
    def _create_resumo_df(self, results: List[ComparisonResult]) -> pd.DataFrame:
        """Cria DataFrame da aba Resumo NF."""
        data = []
        
        for result in results:
            nfe = result.nfe
            
            data.append({
                "Arquivo": nfe.arquivo,
                "Chave NF-e": nfe.chave,
                "Número": nfe.nNF,
                "Série": nfe.serie,
                "Data Emissão": nfe.dhEmi.strftime("%d/%m/%Y %H:%M"),
                "Emitente CNPJ": nfe.emit_CNPJ,
                "Emitente Nome": nfe.emit_xNome,
                "Emitente UF": nfe.emit_UF,
                "Destinatário": nfe.dest_CNPJ or nfe.dest_CPF,
                "Dest. Nome": nfe.dest_xNome,
                "Dest. UF": nfe.dest_UF,
                "Valor NF": float(nfe.vNF),
                "Total Itens": result.total_itens,
                "Base Calc IBS": float(result.total_base_calc_ibs),
                "Base XML IBS": float(result.total_base_xml_ibs),
                "Valor XML IBS": float(result.total_valor_xml_ibs),
                "Base Calc CBS": float(result.total_base_calc_cbs),
                "Base XML CBS": float(result.total_base_xml_cbs),
                "Valor XML CBS": float(result.total_valor_xml_cbs),
                "Diverg. IBS": result.count_divergencias_ibs,
                "Diverg. CBS": result.count_divergencias_cbs,
                "Sem Tag IBS": result.count_sem_tag_ibs,
                "Sem Tag CBS": result.count_sem_tag_cbs,
                "Status": result.status_geral.value,
            })
        
        return pd.DataFrame(data)
    
    def _create_itens_df(self, results: List[ComparisonResult]) -> pd.DataFrame:
        """Cria DataFrame da aba Itens."""
        data = []
        
        for result in results:
            nfe = result.nfe
            
            for item in result.itens_resultado:
                data.append({
                    "Arquivo": nfe.arquivo,
                    "Chave NF-e": nfe.chave,
                    "Número NF": nfe.nNF,
                    "Item": item.nItem,
                    "Cód. Produto": item.cProd,
                    "Descrição": item.xProd,
                    "Base Calc IBS": float(item.base_calc_ibs),
                    "Base XML IBS": float(item.base_xml_ibs) if item.base_xml_ibs else None,
                    "Alíq XML IBS (%)": float(item.aliq_xml_ibs) if item.aliq_xml_ibs else None,
                    "Valor XML IBS": float(item.valor_xml_ibs) if item.valor_xml_ibs else None,
                    "Delta Base IBS": float(item.delta_base_ibs) if item.delta_base_ibs else None,
                    "Delta Valor IBS": float(item.delta_valor_ibs) if item.delta_valor_ibs else None,
                    "Status IBS": item.status_ibs.value,
                    "Diverg Base IBS": "SIM" if item.diverge_base_ibs else "NÃO",
                    "Diverg Valor IBS": "SIM" if item.diverge_valor_ibs else "NÃO",
                    "Base Calc CBS": float(item.base_calc_cbs),
                    "Base XML CBS": float(item.base_xml_cbs) if item.base_xml_cbs else None,
                    "Alíq XML CBS (%)": float(item.aliq_xml_cbs) if item.aliq_xml_cbs else None,
                    "Valor XML CBS": float(item.valor_xml_cbs) if item.valor_xml_cbs else None,
                    "Delta Base CBS": float(item.delta_base_cbs) if item.delta_base_cbs else None,
                    "Delta Valor CBS": float(item.delta_valor_cbs) if item.delta_valor_cbs else None,
                    "Status CBS": item.status_cbs.value,
                    "Diverg Base CBS": "SIM" if item.diverge_base_cbs else "NÃO",
                    "Diverg Valor CBS": "SIM" if item.diverge_valor_cbs else "NÃO",
                    "Status Geral": item.status_geral.value,
                })
        
        return pd.DataFrame(data)
    
    def _create_divergencias_df(self, results: List[ComparisonResult]) -> pd.DataFrame:
        """Cria DataFrame da aba Divergências (somente itens divergentes)."""
        data = []
        
        for result in results:
            nfe = result.nfe
            
            for item in result.itens_resultado:
                # Incluir apenas itens com divergência
                if item.status_geral != ValidationStatus.DIVERGENTE:
                    continue
                
                data.append({
                    "Arquivo": nfe.arquivo,
                    "Chave NF-e": nfe.chave,
                    "Número NF": nfe.nNF,
                    "Emitente": nfe.emit_xNome,
                    "Item": item.nItem,
                    "Cód. Produto": item.cProd,
                    "Descrição": item.xProd,
                    "Base Calc IBS": float(item.base_calc_ibs),
                    "Base XML IBS": float(item.base_xml_ibs) if item.base_xml_ibs else None,
                    "Delta Base IBS": float(item.delta_base_ibs) if item.delta_base_ibs else None,
                    "Diverg Base IBS": "SIM" if item.diverge_base_ibs else "NÃO",
                    "Valor XML IBS": float(item.valor_xml_ibs) if item.valor_xml_ibs else None,
                    "Delta Valor IBS": float(item.delta_valor_ibs) if item.delta_valor_ibs else None,
                    "Diverg Valor IBS": "SIM" if item.diverge_valor_ibs else "NÃO",
                    "Base Calc CBS": float(item.base_calc_cbs),
                    "Base XML CBS": float(item.base_xml_cbs) if item.base_xml_cbs else None,
                    "Delta Base CBS": float(item.delta_base_cbs) if item.delta_base_cbs else None,
                    "Diverg Base CBS": "SIM" if item.diverge_base_cbs else "NÃO",
                    "Valor XML CBS": float(item.valor_xml_cbs) if item.valor_xml_cbs else None,
                    "Delta Valor CBS": float(item.delta_valor_cbs) if item.delta_valor_cbs else None,
                    "Diverg Valor CBS": "SIM" if item.diverge_valor_cbs else "NÃO",
                })
        
        return pd.DataFrame(data)
    
    def _create_erros_df(self, errors: List[ErrorRecord]) -> pd.DataFrame:
        """Cria DataFrame da aba Erros."""
        data = []
        
        for error in errors:
            data.append({
                "Arquivo": error.arquivo,
                "Tipo Erro": error.error_type,
                "Mensagem": error.error_message,
                "Data/Hora": error.timestamp.strftime("%d/%m/%Y %H:%M:%S"),
            })
        
        return pd.DataFrame(data)

    def _create_tags_df(self, results: List[ComparisonResult]) -> pd.DataFrame:
        """Cria DataFrame com todas as tags encontradas em cada NF-e.

        Colunas: Arquivo, Chave NF-e, TagPath, TagName, Valor
        """
        data = []

        import re

        for result in results:
            nfe = result.nfe
            tags = getattr(nfe, 'tags', {}) or {}

            # Se não houver tags, mantemos uma linha indicando ausência
            if not tags:
                data.append({
                    "Arquivo": nfe.arquivo,
                    "Chave NF-e": nfe.chave,
                    "Item": None,
                    "TagPath": "",
                    "TagName": "",
                    "Valor": "",
                })
                continue

            for path, value in tags.items():
                tag_name = path.split('/')[-1] if path else ''
                # extrair número do item se path contiver det[<num>] e mapear para código do produto
                produto_cod = None
                m = re.search(r"det\[(\d+)\]", path or '')
                if m:
                    try:
                        item_num = int(m.group(1))
                        # buscar o produto correspondente em nfe.itens
                        if hasattr(nfe, 'itens'):
                            for item in nfe.itens:
                                if item.nItem == item_num:
                                    produto_cod = item.cProd
                                    break
                    except Exception as e:
                        pass

                data.append({
                    "Arquivo": nfe.arquivo,
                    "Chave NF-e": nfe.chave,
                    "Item": produto_cod if produto_cod else "",
                    "TagPath": path,
                    "TagName": tag_name,
                    "Valor": value if value is not None else "",
                })

        return pd.DataFrame(data)
    
    def _apply_formatting(self):
        """Aplica formatação ao arquivo Excel."""
        wb = load_workbook(self.output_path)
        
        # Formatar cada aba
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            self._format_sheet(ws, sheet_name)
        
        wb.save(self.output_path)
    
    def _format_sheet(self, ws, sheet_name: str):
        """
        Aplica formatação a uma planilha.
        
        Args:
            ws: Worksheet do openpyxl
            sheet_name: Nome da aba
        """
        if ws.max_row == 1:  # Planilha vazia
            return
        
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Freeze header
        ws.freeze_panes = ws["A2"]
        
        # Auto-filter
        ws.auto_filter.ref = ws.dimensions
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Formatação de números (2 casas decimais)
        number_columns = self._get_number_columns(sheet_name)
        for col_letter in number_columns:
            for row in range(2, ws.max_row + 1):
                cell = ws[f"{col_letter}{row}"]
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"
        
        # Destacar divergências na aba de Divergências
        if sheet_name == "Divergências":
            self._highlight_divergencias(ws)
        
        # Destacar status na aba Resumo
        if sheet_name == "Resumo NF":
            self._highlight_status_resumo(ws)

        # Formatação para Combined: destacar linhas separadoras (se existirem)
        if sheet_name == "Combined":
            self._format_combined(ws)

    def _create_combined_sheet(self, writer, results: List[ComparisonResult]):
        """Cria uma aba 'Combined' que contém blocos por documento separados por uma linha de cabeçalho."""
        start_row = 0
        workbook = writer.book
        sheet_name = "Combined"
        # criar sheet vazia para garantir presença
        workbook.create_sheet(title=sheet_name)
        ws = workbook[sheet_name]

        for result in results:
            nfe = result.nfe

            # Cabeçalho separador (arquivo — chave — status)
            header = f"Arquivo: {nfe.arquivo} | Chave: {nfe.chave} | Status: {result.status_geral.value}"
            ws.cell(row=start_row + 1, column=1, value=header)
            ws.cell(row=start_row + 1, column=1).font = Font(bold=True)

            # Informações principais (metadados)
            meta = {
                "Número": nfe.nNF,
                "Série": nfe.serie,
                "Data Emissão": nfe.dhEmi.strftime("%d/%m/%Y %H:%M") if getattr(nfe, 'dhEmi', None) else "",
                "Emitente CNPJ": nfe.emit_CNPJ,
                "Emitente Nome": nfe.emit_xNome,
                "Emitente UF": nfe.emit_UF,
            }
            # Escrever metadados como duas colunas
            r = start_row + 2
            for k, v in meta.items():
                ws.cell(row=r, column=1, value=k)
                ws.cell(row=r, column=2, value=v)
                r += 1

            # Itens (se houver)
            itens_df = None
            if result.itens_resultado:
                itens_df = pd.DataFrame([{
                    "Item": it.nItem,
                    "Cód. Produto": it.cProd,
                    "Descrição": it.xProd,
                    "Base Calc IBS": float(it.base_calc_ibs),
                    "Base XML IBS": float(it.base_xml_ibs) if it.base_xml_ibs else None,
                    "Valor XML IBS": float(it.valor_xml_ibs) if it.valor_xml_ibs else None,
                } for it in result.itens_resultado])

            # Tags (se houver)
            tags = getattr(nfe, 'tags', {}) or {}
            tags_df = None
            if tags:
                # incluir coluna Item para facilitar busca por item no Combined
                import re
                rows = []
                for k, v in tags.items():
                    # extrair código do produto se a tag pertence a det[n]
                    produto_cod = None
                    m = re.search(r"det\[(\d+)\]", k)
                    if m:
                        try:
                            item_num = int(m.group(1))
                            # buscar o produto correspondente
                            if hasattr(nfe, 'itens'):
                                for item in nfe.itens:
                                    if item.nItem == item_num:
                                        produto_cod = item.cProd
                                        break
                        except Exception:
                            pass
                    rows.append({"Item": produto_cod if produto_cod else "", "TagPath": k, "Valor": v})
                tags_df = pd.DataFrame(rows)

            # Escrever itens_df e tags_df usando pandas to_excel com startrow
            write_row = r + 1
            if itens_df is not None and not itens_df.empty:
                ws.cell(row=write_row, column=1, value="Itens")
                ws.cell(row=write_row, column=1).font = Font(bold=True)
                write_row += 1
                itens_df.to_excel(writer, sheet_name=sheet_name, startrow=write_row - 1, index=False)
                # calcular linhas escritas
                write_row += len(itens_df.index) + 1

            if tags_df is not None and not tags_df.empty:
                ws.cell(row=write_row, column=1, value="Tags")
                ws.cell(row=write_row, column=1).font = Font(bold=True)
                write_row += 1
                tags_df.to_excel(writer, sheet_name=sheet_name, startrow=write_row - 1, index=False)
                write_row += len(tags_df.index) + 1

            # Avançar start_row para próximo bloco com uma linha em branco separadora
            start_row = write_row + 2

        # Depois de escrever via pandas, garantir que writer salva as mudanças (pandas fará no exit)

    def _format_combined(self, ws):
        """Formatação simples para a aba Combined: ajustar colunas e destacar headers."""
        # Ajustar largura colunas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 70)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _get_number_columns(self, sheet_name: str) -> List[str]:
        """Retorna letras das colunas numéricas por aba."""
        if sheet_name == "Resumo NF":
            # Colunas: Valor NF, bases, valores, contadores
            return ["L", "N", "O", "P", "Q", "R", "S"]
        elif sheet_name == "Itens":
            # Colunas: bases, alíquotas, valores, deltas
            return ["G", "H", "I", "J", "K", "L", "P", "Q", "R", "S", "T", "U"]
        elif sheet_name == "Divergências":
            # Colunas: bases, deltas, valores
            return ["H", "I", "J", "L", "M", "O", "P", "Q", "S", "T", "U"]
        
        return []
    
    def _highlight_divergencias(self, ws):
        """Destaca células com divergências."""
        divergence_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Procurar coluna "Diverg Base IBS", "Diverg Valor IBS", etc.
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                
                # Se célula contém "SIM" em colunas de divergência
                if cell.value == "SIM":
                    header = ws.cell(row=1, column=col).value
                    if header and "Diverg" in header:
                        cell.fill = divergence_fill
    
    def _highlight_status_resumo(self, ws):
        """Destaca status na aba Resumo."""
        # Encontrar coluna de Status
        status_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "Status":
                status_col = col
                break
        
        if status_col is None:
            return
        
        divergente_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        sem_tag_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=status_col)
            
            if cell.value == "DIVERGENTE":
                cell.fill = divergente_fill
            elif cell.value == "OK":
                cell.fill = ok_fill
            elif cell.value == "SEM_TAG":
                cell.fill = sem_tag_fill
