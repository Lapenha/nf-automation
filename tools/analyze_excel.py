"""Analisa a planilha Excel para extrair tags obrigatórias."""
import openpyxl
from pathlib import Path

# Carrega planilha
excel_path = Path(__file__).parent.parent / "planilhabase" / "Mastersaf_Layout_DFe_V3_NFSe_NACIONAL_1_01 - Oficial Reforma.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['Emissão Nacional NFSe  - V1.01']

print(f"Total de linhas: {ws.max_row}")
print(f"Total de colunas: {ws.max_column}")
print()

# Analisa estrutura
print("Cabeçalhos:")
for col in range(1, 21):
    val = ws.cell(8, col).value
    if val:
        print(f"  Col {col}: {val}")

print("\n\nPrimeiras 30 linhas com dados:")
for i in range(9, 40):
    nome = ws.cell(i, 10).value
    elemento = ws.cell(i, 14).value
    ocorr = ws.cell(i, 16).value
    
    if nome and elemento:
        print(f"Linha {i}: {nome} | {elemento} | {ocorr}")

print("\n\nTags obrigatórias (1-1):")
tags_obrigatorias = []
for i in range(9, ws.max_row + 1):
    nome = ws.cell(i, 10).value
    elemento = ws.cell(i, 14).value
    ocorr = ws.cell(i, 16).value
    
    if nome and elemento and ocorr and '1-1' in str(ocorr):
        tags_obrigatorias.append({
            'nome': nome,
            'elemento': elemento,
            'ocorr': ocorr,
            'linha': i
        })

print(f"\nTotal de tags 1-1: {len(tags_obrigatorias)}")
print("\nPrimeiras 50 tags obrigatórias:")
for tag in tags_obrigatorias[:50]:
    print(f"  {tag['elemento']} ({tag['nome']}) - Linha {tag['linha']}")
