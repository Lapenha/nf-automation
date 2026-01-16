"""
Teste de geração de relatório com Descrição e Observação.
"""
from pathlib import Path
from src.nfe_validator.tag_validator import TagValidator
import openpyxl
from openpyxl.styles import Font, PatternFill

def test_report_with_description():
    """Testa geração de relatório com descrição e observação."""
    
    print("="*80)
    print("TESTE DE RELATÓRIO COM DESCRIÇÃO E OBSERVAÇÃO")
    print("="*80)
    print()
    
    # Caminho da planilha
    base_dir = Path(__file__).parent
    excel_path = base_dir / "planilhabase" / "Mastersaf_Layout_DFe_V3_NFSe_NACIONAL_1_01 - Oficial Reforma.xlsx"
    
    # Cria validador
    print("Carregando validador...")
    validator = TagValidator(excel_path)
    print(f"✅ {len(validator.tags_obrigatorias)} tags carregadas")
    print()
    
    # Testa com um XML
    xml_dir = base_dir / "xmls"
    xml_files = list(xml_dir.glob("*.xml"))[:2]  # Pega apenas 2 para teste
    
    print(f"Validando {len(xml_files)} XML(s)...")
    resultados = validator.validar_multiplos_xmls(xml_files)
    print()
    
    # Gera relatório Excel de teste
    output_path = Path.home() / "Downloads" / "teste_relatorio_tags.xlsx"
    print(f"Gerando relatório em: {output_path}")
    
    wb = openpyxl.Workbook()
    
    # Aba 1: Resumo
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    
    ws_resumo['A1'] = "RELATÓRIO DE VALIDAÇÃO DE TAGS OBRIGATÓRIAS"
    ws_resumo['A1'].font = Font(size=14, bold=True)
    ws_resumo.merge_cells('A1:F1')
    
    ws_resumo['A3'] = "Arquivo XML"
    ws_resumo['B3'] = "Status"
    ws_resumo['C3'] = "Tags Encontradas"
    ws_resumo['D3'] = "Total Tags"
    ws_resumo['E3'] = "Tags Ausentes"
    ws_resumo['F3'] = "% Completude"
    
    for cell in ws_resumo[3]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    linha = 4
    for resultado in resultados:
        ws_resumo[f'A{linha}'] = resultado.xml_path.name
        ws_resumo[f'B{linha}'] = "OK" if resultado.sucesso else "Ausências"
        ws_resumo[f'C{linha}'] = resultado.tags_encontradas
        ws_resumo[f'D{linha}'] = resultado.total_tags_obrigatorias
        ws_resumo[f'E{linha}'] = len(resultado.tags_ausentes)
        
        percentual = (resultado.tags_encontradas / resultado.total_tags_obrigatorias * 100) if resultado.total_tags_obrigatorias > 0 else 0
        ws_resumo[f'F{linha}'] = f"{percentual:.1f}%"
        linha += 1
    
    ws_resumo.column_dimensions['A'].width = 50
    ws_resumo.column_dimensions['B'].width = 15
    ws_resumo.column_dimensions['C'].width = 18
    ws_resumo.column_dimensions['D'].width = 15
    ws_resumo.column_dimensions['E'].width = 18
    ws_resumo.column_dimensions['F'].width = 15
    
    # Aba 2: Tags Ausentes COM DESCRIÇÃO E OBSERVAÇÃO
    ws_detalhes = wb.create_sheet("Tags Ausentes")
    
    ws_detalhes['A1'] = "TAGS OBRIGATÓRIAS AUSENTES POR ARQUIVO"
    ws_detalhes['A1'].font = Font(size=14, bold=True)
    ws_detalhes.merge_cells('A1:G1')
    
    ws_detalhes['A3'] = "Arquivo XML"
    ws_detalhes['B3'] = "Nome da Tag"
    ws_detalhes['C3'] = "Descrição"
    ws_detalhes['D3'] = "Observação"
    ws_detalhes['E3'] = "Elemento"
    ws_detalhes['F3'] = "Tipo"
    ws_detalhes['G3'] = "Linha Planilha"
    
    for cell in ws_detalhes[3]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    linha = 4
    for resultado in resultados:
        if resultado.tags_ausentes:
            # Limita a 10 tags por arquivo para teste
            for tag in resultado.tags_ausentes[:10]:
                ws_detalhes[f'A{linha}'] = resultado.xml_path.name
                ws_detalhes[f'B{linha}'] = tag.nome
                ws_detalhes[f'C{linha}'] = tag.descricao
                ws_detalhes[f'D{linha}'] = tag.observacao
                ws_detalhes[f'E{linha}'] = tag.elemento
                ws_detalhes[f'F{linha}'] = tag.tipo
                ws_detalhes[f'G{linha}'] = tag.linha
                linha += 1
    
    ws_detalhes.column_dimensions['A'].width = 50
    ws_detalhes.column_dimensions['B'].width = 30
    ws_detalhes.column_dimensions['C'].width = 50
    ws_detalhes.column_dimensions['D'].width = 50
    ws_detalhes.column_dimensions['E'].width = 15
    ws_detalhes.column_dimensions['F'].width = 10
    ws_detalhes.column_dimensions['G'].width = 18
    
    # Salva
    wb.save(output_path)
    print(f"✅ Relatório gerado com sucesso!")
    print()
    print(f"Abra o arquivo para conferir:")
    print(f"  {output_path}")
    print()
    print("Verifique na aba 'Tags Ausentes' se as colunas estão presentes:")
    print("  - Nome da Tag")
    print("  - Descrição ← NOVO!")
    print("  - Observação ← NOVO!")
    print("  - Elemento")
    print("  - Tipo")
    print("  - Linha Planilha")
    print()
    
    # Mostra exemplo das primeiras tags ausentes
    if resultados and resultados[0].tags_ausentes:
        print("Exemplo das primeiras 3 tags ausentes:")
        for i, tag in enumerate(resultados[0].tags_ausentes[:3], 1):
            print(f"\n{i}. {tag.nome}")
            print(f"   Descrição: {tag.descricao[:100] if tag.descricao else '(vazio)'}")
            print(f"   Observação: {tag.observacao[:100] if tag.observacao else '(vazio)'}")
    
    print()
    print("="*80)

if __name__ == "__main__":
    test_report_with_description()
